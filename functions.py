#This module contatins all function for a statics problem of a car going up a hill
#contains input, calculation, and output functions
import numpy as np
import openpyxl as pyxl
import matplotlib.pyplot as plt
from numpy.polynomial import Polynomial as poly
workbook= pyxl.load_workbook("data.xlsx")


def getGears(sheet):
    """
    This function takes in the spreadsheet and then uses the sheet to 
    get the different gear ratios, then checks to make sure all values are positive 
    if a value is negative the program prompts the user to change the value to a new value
    
    args:
    sheet: takes in an excell sheet
    returns:
    gear1: 1st gear, gear ratio
    gear2: 2nd gear, gear ratio
    gear3: 3rd gear, gear ratio
    gear4: 4th gear, gear ratio
    gear5: 5th gear, gear ratio
    gear6: 6th gear, gear ratio
    """
    gears=[]
    for i in range(4,10):
        while(sheet["B{}".format(i)].value<0):
            sheet["B{}".format(i)]=float(input("The value of your {} is negative, input a new value: ".format(sheet["A{}".format(i)].value)))
        gears.append(sheet["B{}".format(i)].value)
    gears=np.array(gears)
    return gears
def getothers(sheet):
    """
    this function is being used to grab our data 
    tslope is the the terrain slope as a percentage\
    wbase is the wheel base of the car
    radius is the radius of the cars wheels
    rollress is the rolling resitance coeficiant
    hA is the center of gravity of the car 
    fdrive is the final drive efficiency as a percentage
    teff is the transimion efficiency as a percentage
    weight is the weight of the car given in N
    airden is the airdensity given in ρ
    dratio is the drive ratio of the car 
    centerg is the distance from the center of gravity of the car to the front wheel
    """
    
    for i in range(12,24):
        while(sheet["C{}".format(i)].value<0):
            sheet["C{}".format(i)]=float(input("The value of your {} is negative, input a new value: ".format(sheet["A{}".format(i)].value)))
    
    speed=sheet["C12"].value /3.6
    tslope=sheet["C13"].value/100
    wbase=sheet["C14"].value
    radius=sheet["C15"].value
    rollres=sheet["C16"].value
    hA=sheet["C17"].value
    fdrive=sheet["C18"].value
    teff=sheet["C19"].value
    weight=sheet["C20"].value
    airden=sheet["C21"].value
    dratio=sheet["C22"].value
    centerg=sheet["C23"].value
    return speed,tslope,wbase,radius,rollres,hA,fdrive,teff,weight,airden,dratio,centerg

def getdyno(sheet):
    """this function extracts the data from our excel file from our dyno data that we collected"""
    angularvex=[]
    torquex=[]
    for i in range(4,16):
        angularvex.append(sheet["A{}".format(i)].value)
        
        torquex.append(sheet["B{}".format(i)].value)
    angularvex=np.array(angularvex)
    torquex=np.array(torquex)
    
    return angularvex,torquex

def doMath(rollres,weight,tslope,airden,dragC,csA,v,radius,dratio,teff,fdrive,gears,angularvex,torquex,name,wbase,centerg,hA):
    """
    This is the function that does math
    airress calculates airressitance that is affecting the car
    rRoll is the caculation for the rolling ressitance 
    roadload is the caculation of rolling restiance + the grade load + the airres
    angularve is the caculation of angualrvelocity derived from gears * the driveratio * velocity * the wheel area
    finaledrivE is fdrive converted from a percentage 
    the if function is to differntiate between the data from the cars given and the car we are getting new dyno data from
    te is the torque we get from the best poly fit of angularvex and the angularve
    torqued is the caculation of the torque at the drive train using finaledrivE and the drive ratio from the data provide
    teff is the percentage of transmision efficiany given which is then converted from a percentage
    traction is the caculation of torqued over the radius of the car wheel
    wperp is a shorthand variable for caculating weight * cos theta

    """
    airres=.5*airden*dragC*csA*v**2
    rRoll=rollres*weight*np.cos((np.atan(tslope)))
    roadLoad=rRoll+weight*np.sin((np.atan(tslope)))+airres


    angularve=gears*dratio*v*60/(radius*2*np.pi)
    finaldrivE=fdrive/100
    if(name!="CR-28"):
        fx=poly.fit(angularvex,torquex,4)
    else:
        fx=poly.fit(angularvex,torquex,3)
    te=fx(angularve)
    torqued=finaldrivE*dratio*(teff/100)*gears*te
    traction=torqued/radius
    acceleration=((((traction-roadLoad))*9.81)/weight)
    Wperp=weight*np.cos(tslope)
    
    
    frontload=np.absolute(((airres*hA)+(weight/9.81)*hA*acceleration-(wbase-centerg)*Wperp+weight*np.sin(tslope)*hA)/wbase)
    rearload=((airres*hA)+(weight/9.81)*hA*acceleration+(centerg)*Wperp+weight*np.sin(tslope)*hA)/wbase

    frontloads=(Wperp*(wbase-centerg))/(wbase)
    rearloads=(Wperp*centerg)/wbase


    return traction,torqued,acceleration, angularve, roadLoad,rearload,frontload,frontloads,rearloads,te



def calchp(angularve,torque):
    """
    calculates horsepower given torque and angular velocity
    args:
    torque= given torque
    angularve= omega given
    return
    hp= horsepower
    """
    hp=torque*angularve*0.7375621493/5252
    return hp
    
def graphs(rollres,weight,tslope,airden,dragC,csA,v,radius,dratio,teff,fdrive,gears,angularvex,torquex,name,wbase,centerg,hA):
    #dyno graph
    if(name!="CR-28"):
        fx=poly.fit(angularvex,torquex,4)
    else:
        fx=poly.fit(angularvex,torquex,3)
    x=np.linspace(0,max(angularvex),6000)
    fy=fx(x)

    plt.subplot(2,2,1)
    plt.plot(angularvex,torquex,"ro")
    plt.plot(x,fy,"b-")
    plt.title("Experimental dyno data and fit")
    plt.ylabel("engine torque(n m)")
    plt.xlabel("angular velocity(rpm)")
    plt.legend([name])
    if(name=="CR-28"):
        plt.ylim(0,34)
    colors=["b","c","g","k","r","m"]
    vels=np.linspace(0,200,200)
    rads=np.linspace(.1,.5,60)
    weigh=np.linspace(1000,20000,100)
    for n,each_gear in enumerate(gears):
        for i,each_vel in enumerate(vels):
           plt.subplot(2,2,2)
           
           traction,torqued,acceleration, angularve, roadLoad,rearload,frontload,frontloads,rearloads,te= doMath(rollres,weight,tslope,airden,dragC,csA,each_vel,radius,dratio,teff,fdrive,each_gear,angularvex,torquex,name,wbase,centerg,hA)
           if angularve<7000 :
                
                plt.plot(each_vel,acceleration,"{}o".format(colors[n]),markersize=2)
                
                   
                if i==5:
                    plt.text(each_vel+0.05,acceleration+0.05,"gear {}".format(n+1),fontsize=8)
        plt.title("velocity vs acceleration")
        plt.xlabel("velocity (m/s)")
        plt.ylabel("Acceleration (m/s^2)") 
        plt.subplot(2,2,3)
        for i,each_rad in enumerate(rads):
            traction,torqued,acceleration, angularve, roadLoad,rearload,frontload,frontloads,rearloads,te= doMath(rollres,weight,tslope,airden,dragC,csA,v,each_rad,dratio,teff,fdrive,each_gear,angularvex,torquex,name,wbase,centerg,hA)
            if angularve<7000 :
                plt.plot(each_rad,traction,"{}o".format(colors[n]),markersize=1)
                if i==40:
                    plt.text(each_rad+0.05,traction-0.05,"gear {}".format(n+1),fontsize=8)
        
        plt.title("radius of wheel vs traction")
        plt.xlabel("radius of wheel (m)")
        plt.ylabel("traction(n)")

        plt.subplot(2,2,4)
        for i,each_weight in enumerate(weigh):
            traction,torqued,acceleration, angularve, roadLoad,rearload,frontload,frontloads,rearloads,te= doMath(rollres,each_weight,tslope,airden,dragC,csA,v,radius,dratio,teff,fdrive,each_gear,angularvex,torquex,name,wbase,centerg,hA)
            if angularve<7000 :
                plt.plot(each_weight,acceleration,"{}o".format(colors[n]),markersize=1)
                if i==10:
                    plt.text(each_weight+0.05,acceleration+0.05,"gear {}".format(n+1),fontsize=8)
        plt.title("weight vs acceleration")
        plt.xlabel("weight(N)")
        plt.ylabel("acceleration(m/s^2)")
        


    plt.tight_layout()
    plt.show()
    
    #



def makeres(count,ogsheet,dragC,area):
    """
    This creates a new workbook and sheet for the results page
    Args:
    count= how many times the user has gone through the loop
    ogsheet= original data sheet
    dragC=drag coefficient
    area=area of the front of vehicle 
    returns:
    newsheet= sheet that would contain new results
    results= workbook that contains results

    """
    try:
        #tests to see if file exists
        results=pyxl.load_workbook("All_Results.xlsx")
    except FileNotFoundError:
        #creates new workbook if file does not exist
        results=pyxl.Workbook()
    
    try:
        #tests to see if sheet exists
        newsheet=results["Results_{}".format(count)]
    except KeyError:
        #creates sheet if the sheet does not exist
        
        newsheet=results.create_sheet("Results_{}".format(count))
    #prints all predetermined info
    newsheet["A4"]="Gears"
    newsheet["B4"]="Ratio"
    newsheet["C4"]="WR(N)"
    newsheet["D4"]="WF(N)"
    newsheet["E4"]="WRstat(N)"
    newsheet["F4"]="WFstat(N)"
    newsheet["G4"]="a(m/s^2)"
    newsheet["H4"]="P(N)"
    newsheet["I4"]="R(N)"
    newsheet["J4"]="HP(hp)"
    newsheet["K4"]="Te(Nm)"
    newsheet["L4"]="omega_e(RPM)"
    newsheet["A12"]="DATA"
    for i in range(12,24):
        newsheet["A{}".format(i)]=ogsheet["B{}".format(i)].value
        newsheet["B{}".format(i)]=ogsheet["C{}".format(i)].value
    newsheet["A24"]="CD"
    newsheet["B24"]=dragC
    newsheet["A25"]="Area"
    newsheet["B25"]=area
    newsheet["A5"]="1st"
    newsheet["A6"]="2nd"
    newsheet["A7"]="3rd"
    newsheet["A8"]="4th"
    newsheet["A9"]="5th"
    newsheet["A10"]="6th"

    results.save("All_Results.xlsx")
    return newsheet,results

def outputs(gears,angularve,te,accel,trac,roadload,newsheet,name,rearloadS,frontLoadS,frontLoad,rearLoad,results):
    """
    outputs all calculated data and given data onto a result exccel file
    args:
    gears: array containing the gear ratio for each gear
    angularve: angular velocity of engine
    te: torque at engine
    accel= acceleration
    trac= traction
    roadload= all resistances and grade load
    newsheet= sheet for data output
    name= name of car
    rearloadS= static rear axle load
    frontLoadS= static front axle load
    frontLoad= front dynamic axle load
    rearLoad= rear dynamic axle load
    results= result workbook
    returns:
    None
    
    """
    for i in range(0,6):
        hp=calchp(angularve[i],te[i])
        #errors here
        if(angularve[i]>7000 and name!="CR-28"):
            newsheet["E{}".format(i+5)]="OOR"
            newsheet["F{}".format(i+5)]="OOR"
            newsheet["G{}".format(i+5)]="OOR"
            newsheet["H{}".format(i+5)]="OOR"
            newsheet["I{}".format(i+5)]="OOR"
            newsheet["J{}".format(i+5)]="OOR"
            newsheet["B{}".format(i+5)]=gears[i]
            newsheet["K{}".format(i+5)]="OOR"
            newsheet["L{}".format(i+5)]="OOR"
            newsheet["C{}".format(i+5)]="OOR"
            newsheet["D{}".format(i+5)]="OOR"
        else:
            
            newsheet["B{}".format(i+5)]=gears[i]
            newsheet["E{}".format(i+5)]=rearloadS
            newsheet["F{}".format(i+5)]=frontLoadS
            newsheet["G{}".format(i+5)]=accel[i]
            newsheet["H{}".format(i+5)]=trac[i]
            newsheet["I{}".format(i+5)]=roadload
            newsheet["J{}".format(i+5)]=hp
        
            newsheet["K{}".format(i+5)]=te[i]
            newsheet["L{}".format(i+5)]=angularve[i]
            newsheet["C{}".format(i+5)]=rearLoad[i]
            newsheet["D{}".format(i+5)]=frontLoad[i]
        results.save("All_Results.xlsx")
def carChoice(workbook):
    """
    This function gives the user an option for which car they want
    and then returns the drag coefficient and area of front bumbper
    Args:
    sheet= needs a sheet to get all value
    Returns:
    dragC= Drag coefficient
    area= area of front hood
    
    
    """
    count=0
    option=0
    sheet1=workbook["Problem Variables"]
    sheet2=workbook["cr-28 var"]
    print("**********************************************")
    print("1. {} {}\n2. {} {}\n3. {} {}\n4.{} {}\n5. {} {}\n6. {} {}\n7.Exit".format(sheet1["B26"].value,sheet1["A26"].value,sheet1["B27"].value,sheet1["A27"].value, sheet1["B28"].value,sheet1["A28"].value,sheet1["B29"].value,sheet1["A29"].value,sheet1["B30"].value,sheet1["A30"].value,sheet2["B26"].value,sheet2["A26"].value))
    print("**********************************************")
    
    option=0
    while(option!=7 and option!= 6 and option!=5 and option!=4 and option!=3 and option!=2 and option!=1):
        
        option=int(input("Select the car choice options (1-7) "))
        match option:
            case 1:
                dragC=sheet1["C26"].value
                area=sheet1["D26"].value
                name=sheet1["A26"].value
                count+=1
                sheet=sheet1

            case 2:
                dragC=sheet1["C27"].value
                area=sheet1["D27"].value
                name=sheet1["A27"].value
                count+=1
                sheet=sheet1
            case 3:
                dragC=sheet1["C28"].value
                area=sheet1["D28"].value
                name=sheet1["A28"].value
                count+=1
                sheet=sheet1
            case 4:
                dragC=sheet1["C29"].value
                area=sheet1["D29"].value
                name=sheet1["A29"].value
                count+=1
                sheet=sheet1
            case 5:
                dragC=sheet1["C30"].value
                area=sheet1["D30"].value
                name=sheet1["A30"].value
                count+=1
                sheet=sheet1
            case 6: 
            
                dragC=sheet2["C26"].value
                area=sheet2["D26"].value
                name=sheet2["A26"].value
                count+=1
                sheet=sheet2
            case 7:
                print("Have a good day")
                name="null"
                area=-1
                dragC=-1
                sheet=sheet1

            case _:
            #HR Friendly error message
                print("ERROR! Selection is invalid\n")
    return dragC,area,name,count,sheet,option

def main(workbook):
        """
        This is the main function for our program
        it gather up our other functions and runs them
        it has a while loop to ask for which car the user would like to have 
        graphs is the function that runs our graphs and outputs them
        makeresis the function that creates a new workbook
        outputs is the function that outputs our data to the excel file 
        """
        option=0
        while(option!=7):
            dragC,area,name,count,sheet,option=carChoice(workbook)
            if(option!=7):
                gears = getGears(sheet)
                
                #   Function that returns all other varibles in the excel for use by later functions
            
                v,tslope,wbase,radius,rollres,hA,fdrive,teff,weight,airden,dratio,centerg = getothers(sheet)
                
                # Function that returns angularvex and torquex
                if(name!="CR-28"):
                    dynosheet=workbook["Dynamometer"]
                else:
                    
                    dynosheet=workbook["cr-dyno"]
            
                angularvex,torquex = getdyno(dynosheet)
            
                #   Function that calculates roadload given rollres,weight,tslope,airden,dragC,csA, and v
                traction,torqued,acceleration, angularve, roadLoad,rearload,frontload,frontloads,rearloads,te=doMath(rollres,weight,tslope,airden,dragC,area,v,radius,dratio,teff,fdrive,gears,angularvex,torquex,name,wbase,centerg,hA)
            
                #   Outputs graphs
                #functions.graphs(angularve, torqued,name)
            
                graphs(rollres,weight,tslope,airden,dragC,area,v,radius,dratio,teff,fdrive,gears,angularvex,torquex,name,wbase,centerg,hA)
                newsheet,results=makeres(count,sheet,dragC,area)    
                outputs(gears,angularve,te,acceleration,traction,roadLoad,newsheet,name,rearloads,frontloads,frontload,rearload,results)
            
